"""Shared helpers for office document tools."""
from __future__ import annotations

import csv
import html.parser
import re
from pathlib import Path
from typing import Iterable


def resolve_path(cwd: str | None, value: str) -> Path:
    return (Path(cwd or ".") / value).resolve()


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


class TextHTMLParser(html.parser.HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag in {"p", "div", "section", "article", "br", "li", "tr", "h1", "h2", "h3"}:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        value = data.strip()
        if value:
            self.parts.append(value)

    def text(self) -> str:
        return "\n".join(part.strip() for part in self.parts if part.strip())


def html_to_text(value: str) -> str:
    parser = TextHTMLParser()
    parser.feed(value)
    return parser.text()


def markdown_blocks(value: str) -> list[tuple[str, str]]:
    blocks: list[tuple[str, str]] = []
    in_code = False
    code_lines: list[str] = []
    paragraph: list[str] = []

    def flush_paragraph() -> None:
        if paragraph:
            blocks.append(("paragraph", " ".join(paragraph).strip()))
            paragraph.clear()

    for raw in value.splitlines():
        line = raw.rstrip()
        stripped = line.strip()
        if stripped.startswith("```"):
            if in_code:
                blocks.append(("code", "\n".join(code_lines)))
                code_lines.clear()
                in_code = False
            else:
                flush_paragraph()
                in_code = True
            continue
        if in_code:
            code_lines.append(line)
            continue
        if not stripped:
            flush_paragraph()
            continue
        if stripped.startswith("#"):
            flush_paragraph()
            level = len(stripped) - len(stripped.lstrip("#"))
            blocks.append((f"h{min(level, 3)}", stripped[level:].strip()))
        elif stripped.startswith(("- ", "* ")):
            flush_paragraph()
            blocks.append(("bullet", stripped[2:].strip()))
        else:
            paragraph.append(stripped)
    flush_paragraph()
    if code_lines:
        blocks.append(("code", "\n".join(code_lines)))
    return blocks


def plain_blocks(value: str, source_format: str) -> list[tuple[str, str]]:
    if source_format == "markdown":
        return markdown_blocks(value)
    if source_format == "html":
        return markdown_blocks(html_to_text(value))
    return [("paragraph", line.strip()) for line in value.splitlines() if line.strip()]


def read_table(path: Path, *, sheet: str | None = None, max_rows: int = 500) -> list[list[str]]:
    if path.suffix.lower() in {".csv", ".tsv"}:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
        with path.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f, delimiter=delimiter)
            return [[str(cell) for cell in row] for _, row in zip(range(max_rows), reader)]

    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover - dependency guidance
            raise RuntimeError("openpyxl is required to read XLSX files") from exc

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        rows = []
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx >= max_rows:
                break
            rows.append(["" if value is None else str(value) for value in row])
        wb.close()
        return rows

    raise ValueError(f"Unsupported table file: {path.suffix}")


def sanitize_filename(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9._-]+", "-", value.strip())
    return value.strip("-") or "asset"


def table_to_markdown(rows: Iterable[Iterable[str]]) -> str:
    rows_list = [list(row) for row in rows]
    if not rows_list:
        return ""
    header = rows_list[0]
    sep = ["---"] * len(header)
    body = rows_list[1:]
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(sep) + " |",
    ]
    for row in body:
        padded = row + [""] * (len(header) - len(row))
        lines.append("| " + " | ".join(padded[: len(header)]) + " |")
    return "\n".join(lines)
